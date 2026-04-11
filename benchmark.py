"""
benchmark.py

Mede o desempenho do pipeline AIAA para diferentes modelos LLM,
com múltiplas execuções por pergunta, e guarda os resultados em
CSV + Excel com estatísticas e gráficos.

Mede três fases:
    1. Embedding + Retrieval  — tempo de pesquisa no ChromaDB
    2. LLM Generation         — tempo de geração da resposta
    3. Total                  — tempo ponta a ponta

Uso:
    # Benchmark com configuração padrão (10 execuções, modelos do config)
    python benchmark.py

    # 50 execuções, modelos específicos
    python benchmark.py --runs 50 --models llama3.2 qwen2.5:7b mistral

    # Só retrieval (sem LLM) — mais rápido
    python benchmark.py --no-llm --runs 50
"""

import argparse
import csv
import json
import os
import statistics
import time
from datetime import datetime

# ---------------------------------------------------------------------------
# Perguntas de benchmark — cobrindo diferentes tipos de recuperação
# ---------------------------------------------------------------------------

PERGUNTAS_BENCHMARK = [
    "Quais os pressupostos da responsabilidade civil extracontratual?",
    "Qual o prazo de prescrição do direito à indemnização?",
    "O que constitui justa causa de despedimento?",
    "Qual o prazo para instaurar procedimento disciplinar?",
    "Quais os efeitos do despedimento ilícito?",
    "O que são danos não patrimoniais e quando são indemnizáveis?",
    "Qual a diferença entre dolo e mera culpa na responsabilidade civil?",
    "Quais as formas de cessação do contrato de trabalho?",
]


# ---------------------------------------------------------------------------
# Timing helpers
# ---------------------------------------------------------------------------

class MedicaoTempo:
    """Regista uma medição individual."""
    def __init__(self, modelo, pergunta, fase, duracao_ms, n_chunks=0, score_top=0.0, erro=None):
        self.timestamp    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.modelo       = modelo
        self.pergunta     = pergunta[:60]
        self.fase         = fase
        self.duracao_ms   = round(duracao_ms, 1)
        self.n_chunks     = n_chunks
        self.score_top    = round(score_top, 3)
        self.erro         = erro or ""

    def to_dict(self):
        return {
            "timestamp":   self.timestamp,
            "modelo":      self.modelo,
            "pergunta":    self.pergunta,
            "fase":        self.fase,
            "duracao_ms":  self.duracao_ms,
            "n_chunks":    self.n_chunks,
            "score_top":   self.score_top,
            "erro":        self.erro,
        }


CSV_FIELDS = ["timestamp","modelo","pergunta","fase","duracao_ms","n_chunks","score_top","erro"]


def guardar_csv(medicoes: list, caminho: str):
    escrever_header = not os.path.exists(caminho)
    with open(caminho, "a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        if escrever_header:
            w.writeheader()
        for m in medicoes:
            w.writerow(m.to_dict())


# ---------------------------------------------------------------------------
# Benchmark core
# ---------------------------------------------------------------------------

def benchmark_retrieval(store, pergunta: str, n_results: int) -> MedicaoTempo:
    """Mede só o tempo de embedding + pesquisa no ChromaDB."""
    t0 = time.perf_counter()
    resultados = store.search(pergunta, n=n_results)
    duracao_ms = (time.perf_counter() - t0) * 1000

    score_top = resultados[0]["score"] if resultados else 0.0
    return MedicaoTempo(
        modelo    = "retrieval_only",
        pergunta  = pergunta,
        fase      = "retrieval",
        duracao_ms= duracao_ms,
        n_chunks  = len(resultados),
        score_top = score_top,
    )


def benchmark_llm(retriever, pergunta: str, modelo: str) -> tuple:
    """
    Mede retrieval e LLM separadamente.
    Retorna (medicao_retrieval, medicao_llm, medicao_total)
    """
    from prompt_builder import PromptBuilder

    # --- Retrieval ---
    t0 = time.perf_counter()
    chunks = retriever._store.search(pergunta, n=retriever._n_results)
    t_retrieval = (time.perf_counter() - t0) * 1000

    score_top = chunks[0]["score"] if chunks else 0.0

    med_retrieval = MedicaoTempo(
        modelo    = modelo,
        pergunta  = pergunta,
        fase      = "retrieval",
        duracao_ms= t_retrieval,
        n_chunks  = len(chunks),
        score_top = score_top,
    )

    # --- LLM ---
    prompt = retriever._prompt_builder.build(pergunta, chunks)
    erro   = None
    t0     = time.perf_counter()
    try:
        retriever._llm.generate(prompt)
    except Exception as e:
        erro = str(e)[:100]
    t_llm = (time.perf_counter() - t0) * 1000

    med_llm = MedicaoTempo(
        modelo    = modelo,
        pergunta  = pergunta,
        fase      = "llm",
        duracao_ms= t_llm,
        n_chunks  = len(chunks),
        score_top = score_top,
        erro      = erro,
    )

    med_total = MedicaoTempo(
        modelo    = modelo,
        pergunta  = pergunta,
        fase      = "total",
        duracao_ms= t_retrieval + t_llm,
        n_chunks  = len(chunks),
        score_top = score_top,
        erro      = erro,
    )

    return med_retrieval, med_llm, med_total


# ---------------------------------------------------------------------------
# Estatísticas
# ---------------------------------------------------------------------------

def calcular_stats(valores: list) -> dict:
    if not valores:
        return {}
    return {
        "n":        len(valores),
        "media":    round(statistics.mean(valores), 1),
        "mediana":  round(statistics.median(valores), 1),
        "desvio":   round(statistics.stdev(valores), 1) if len(valores) > 1 else 0.0,
        "minimo":   round(min(valores), 1),
        "maximo":   round(max(valores), 1),
        "p95":      round(sorted(valores)[int(len(valores)*0.95)], 1) if len(valores) >= 20 else round(max(valores), 1),
    }


def resumo_por_modelo(medicoes: list) -> dict:
    """Agrupa medições por modelo e fase e calcula estatísticas."""
    grupos = {}
    for m in medicoes:
        chave = (m.modelo, m.fase)
        if chave not in grupos:
            grupos[chave] = []
        if not m.erro:
            grupos[chave].append(m.duracao_ms)

    resumo = {}
    for (modelo, fase), valores in grupos.items():
        if modelo not in resumo:
            resumo[modelo] = {}
        resumo[modelo][fase] = calcular_stats(valores)
    return resumo


def imprimir_resumo(resumo: dict):
    print("\n" + "=" * 70)
    print("RESUMO DE BENCHMARK — AIAA Pipeline")
    print("=" * 70)
    for modelo, fases in resumo.items():
        print(f"\n  Modelo: {modelo}")
        print(f"  {'Fase':<12} {'N':>4} {'Média':>8} {'Mediana':>8} {'Desvio':>8} {'Mín':>8} {'Máx':>8} {'P95':>8}")
        print(f"  {'-'*70}")
        for fase in ["retrieval", "llm", "total"]:
            if fase not in fases:
                continue
            s = fases[fase]
            print(f"  {fase:<12} {s['n']:>4} {s['media']:>7.0f}ms {s['mediana']:>7.0f}ms "
                  f"{s['desvio']:>7.0f}ms {s['minimo']:>7.0f}ms {s['maximo']:>7.0f}ms {s['p95']:>7.0f}ms")
    print()


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def gerar_excel(medicoes: list, resumo: dict, caminho: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl não instalado — a saltar geração de Excel")
        print("  Instala com: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    accent  = "1A3A5C"
    green   = "1E5C3A"
    lt_blue = "E8EEF5"
    lt_grey = "F7F5F0"

    def header_style(cell, bg=accent):
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def data_style(cell, bg="FFFFFF"):
        cell.font      = Font(name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="right", vertical="center")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Sheet 1: Dados brutos ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Dados Brutos"

    headers = ["Timestamp","Modelo","Pergunta","Fase","Duração (ms)","N Chunks","Score Top","Erro"]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        header_style(c)
        c.border = border

    for row_i, m in enumerate(medicoes, 2):
        vals = [m.timestamp, m.modelo, m.pergunta, m.fase,
                m.duracao_ms, m.n_chunks, m.score_top, m.erro]
        bg = lt_blue if row_i % 2 == 0 else "FFFFFF"
        for col, v in enumerate(vals, 1):
            c = ws1.cell(row=row_i, column=col, value=v)
            c.font      = Font(name="Arial", size=10)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left" if col <= 4 else "right", vertical="center")
            c.border    = border

    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 45
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 10
    ws1.column_dimensions["G"].width = 10
    ws1.column_dimensions["H"].width = 30
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Resumo estatístico ──────────────────────────────────────
    ws2 = wb.create_sheet("Resumo Estatístico")

    stat_headers = ["Modelo","Fase","N","Média (ms)","Mediana (ms)",
                    "Desvio Padrão (ms)","Mínimo (ms)","Máximo (ms)","P95 (ms)"]
    for col, h in enumerate(stat_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        header_style(c)
        c.border = border

    row_i = 2
    for modelo, fases in resumo.items():
        for fase in ["retrieval", "llm", "total"]:
            if fase not in fases:
                continue
            s  = fases[fase]
            bg = lt_blue if row_i % 2 == 0 else "FFFFFF"
            vals = [modelo, fase, s["n"], s["media"], s["mediana"],
                    s["desvio"], s["minimo"], s["maximo"], s["p95"]]
            for col, v in enumerate(vals, 1):
                c = ws2.cell(row=row_i, column=col, value=v)
                c.font      = Font(name="Arial", size=10)
                c.fill      = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="left" if col <= 2 else "right")
                c.border    = border
                if col >= 4:
                    c.number_format = "#,##0.0"
            row_i += 1

    for col, width in enumerate([28,12,6,14,14,18,12,12,10], 1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Gráfico de médias por modelo ────────────────────────────
    ws3 = wb.create_sheet("Gráfico Médias")

    # Build data table for chart
    modelos  = list(resumo.keys())
    ws3["A1"] = "Modelo"
    ws3["B1"] = "Retrieval (ms)"
    ws3["C1"] = "LLM (ms)"
    ws3["D1"] = "Total (ms)"
    for c in [ws3["A1"], ws3["B1"], ws3["C1"], ws3["D1"]]:
        header_style(c)

    for i, modelo in enumerate(modelos, 2):
        ws3.cell(row=i, column=1, value=modelo).font = Font(name="Arial", size=10)
        for j, fase in enumerate(["retrieval","llm","total"], 2):
            val = resumo[modelo].get(fase, {}).get("media", 0)
            c   = ws3.cell(row=i, column=j, value=val)
            c.font         = Font(name="Arial", size=10)
            c.number_format = "#,##0.0"

    if len(modelos) >= 1:
        chart = BarChart()
        chart.type    = "col"
        chart.title   = "Tempo Médio por Modelo e Fase (ms)"
        chart.y_axis.title = "Tempo (ms)"
        chart.x_axis.title = "Modelo"
        chart.width   = 22
        chart.height  = 14

        data   = Reference(ws3, min_col=2, max_col=4, min_row=1, max_row=1+len(modelos))
        cats   = Reference(ws3, min_col=1, min_row=2, max_row=1+len(modelos))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws3.add_chart(chart, "F2")

    wb.save(caminho)
    print(f"  Excel guardado: {caminho}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="AIAA Benchmark")
    parser.add_argument("--runs",    type=int,  default=10,
                        help="Número de execuções por pergunta (default: 10)")
    parser.add_argument("--models",  nargs="+", default=None,
                        help="Modelos Ollama a testar (default: usa config.json)")
    parser.add_argument("--no-llm",  action="store_true",
                        help="Só mede retrieval, sem chamar o LLM")
    parser.add_argument("--output",  default="benchmark_resultados",
                        help="Prefixo dos ficheiros de saída (default: benchmark_resultados)")
    args = parser.parse_args()

    from config import load_config, build_embedder, build_store, \
                       build_prompt_builder, build_llm
    from retriever import Retriever

    cfg      = load_config()
    embedder = build_embedder(cfg.embedder)
    store    = build_store(cfg.vector_store, embedder)
    builder  = build_prompt_builder(cfg.prompt_builder)

    if store.count() == 0:
        print("ERRO: store vazio. Corre primeiro: python indexar.py")
        return

    # Modelos a testar
    if args.no_llm:
        modelos = ["retrieval_only"]
    else:
        modelos = args.models or [cfg.llm.model_name]

    csv_path   = f"{args.output}.csv"
    excel_path = f"{args.output}.xlsx"
    todas_medicoes = []

    print(f"\nAIAA Benchmark")
    print(f"  Execuções por pergunta : {args.runs}")
    print(f"  Perguntas              : {len(PERGUNTAS_BENCHMARK)}")
    print(f"  Modelos                : {modelos}")
    print(f"  Total de medições      : {args.runs * len(PERGUNTAS_BENCHMARK) * len(modelos)}")
    print(f"  Ficheiro CSV           : {csv_path}")
    print(f"  Ficheiro Excel         : {excel_path}\n")

    for modelo in modelos:
        print(f"\n{'='*60}")
        print(f"  Modelo: {modelo}")
        print(f"{'='*60}")

        # Criar retriever com este modelo
        if not args.no_llm:
            cfg.llm.model_name = modelo
            llm       = build_llm(cfg.llm)
            retriever = Retriever(store, builder, llm, cfg.retriever.n_results)

        medicoes_modelo = []

        for pergunta in PERGUNTAS_BENCHMARK:
            print(f"\n  Pergunta: {pergunta[:55]}...")
            print(f"  {'Run':>4}  {'Retrieval':>10}  {'LLM':>10}  {'Total':>10}")
            print(f"  {'-'*40}")

            for run in range(1, args.runs + 1):
                if args.no_llm:
                    m = benchmark_retrieval(store, pergunta, cfg.retriever.n_results)
                    medicoes_modelo.append(m)
                    if run % 10 == 0 or run == args.runs:
                        print(f"  {run:>4}  {m.duracao_ms:>9.0f}ms")
                else:
                    mr, ml, mt = benchmark_llm(retriever, pergunta, modelo)
                    medicoes_modelo.extend([mr, ml, mt])
                    status = "✓" if not mt.erro else "✗"
                    print(f"  {run:>4}  {mr.duracao_ms:>9.0f}ms  {ml.duracao_ms:>9.0f}ms  {mt.duracao_ms:>9.0f}ms  {status}")

        # Guardar CSV incrementalmente
        guardar_csv(medicoes_modelo, csv_path)
        todas_medicoes.extend(medicoes_modelo)
        print(f"\n  ✓ {len(medicoes_modelo)} medições guardadas no CSV")

    # Calcular e imprimir resumo
    resumo = resumo_por_modelo(todas_medicoes)
    imprimir_resumo(resumo)

    # Gerar Excel
    gerar_excel(todas_medicoes, resumo, excel_path)

    print(f"\nFicheiros gerados:")
    print(f"  {csv_path}   — dados brutos (todas as medições)")
    print(f"  {excel_path} — resumo estatístico + gráfico\n")


if __name__ == "__main__":
    main()